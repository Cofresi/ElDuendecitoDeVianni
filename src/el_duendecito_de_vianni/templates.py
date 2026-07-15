from __future__ import annotations

import logging
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.cell import range_boundaries

PLACEHOLDER_RE = re.compile(r"\{\{([^{}]+)\}\}")
STATIC_TEMPLATE_MARKER = "__static__"


@dataclass
class TemplateResult:
    generated_files: list[Path] = field(default_factory=list)
    missing_placeholders: set[str] = field(default_factory=set)
    skipped_files: list[str] = field(default_factory=list)


def placeholders_in_text(text: str) -> set[str]:
    return set(PLACEHOLDER_RE.findall(text or ""))


def _parse_placeholder(placeholder: str) -> tuple[str, str]:
    key, separator, formatter = placeholder.partition("|")
    return key.strip(), formatter.strip().casefold() if separator else ""


def _decimal_from_text(value: str) -> Decimal | None:
    text = str(value or "").strip()
    if not text:
        return None
    cleaned = text.replace(",", "")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _format_money(value: str) -> str:
    number = _decimal_from_text(value)
    if number is None:
        return ""
    return f"{number:,.2f}"


def _format_int(value: str) -> str:
    number = _decimal_from_text(value)
    if number is None:
        return ""
    return str(int(number))


def _format_date(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return text


def _format_treatment(value: str) -> str:
    text = str(value or "").strip().casefold()
    if text in {"f", "fem", "femenino", "mujer"}:
        return "Sra."
    if text in {"m", "masc", "masculino", "hombre"}:
        return "Sr."
    return ""


def _format_article(value: str, lowercase: bool = False) -> str:
    gender = _gender_code(value)
    if gender == "f":
        return "la" if lowercase else "La"
    if gender == "m":
        return "el" if lowercase else "El"
    return ""


def _gender_code(value: str) -> str:
    text = str(value or "").strip().casefold()
    if text in {"f", "fem", "femenino", "mujer"}:
        return "f"
    if text in {"m", "masc", "masculino", "hombre"}:
        return "m"
    return ""


def _format_gender_ending(value: str, plural: bool = False) -> str:
    gender = _gender_code(value)
    if gender == "f":
        return "as" if plural else "a"
    if gender == "m":
        return "os" if plural else "o"
    return ""


def _format_gender_noun_ending(value: str, plural: bool = False) -> str:
    gender = _gender_code(value)
    if gender == "f":
        return "as" if plural else "a"
    if gender == "m":
        return "es" if plural else ""
    return ""


def _format_placeholder_value(value: str, formatter: str) -> str:
    if not formatter:
        return value
    if formatter == "money":
        return _format_money(value)
    if formatter == "int":
        return _format_int(value)
    if formatter == "date":
        return _format_date(value)
    if formatter == "tratamiento":
        return _format_treatment(value)
    if formatter == "articulo":
        return _format_article(value)
    if formatter == "articulo_minuscula":
        return _format_article(value, lowercase=True)
    if formatter == "genero":
        return _format_gender_ending(value)
    if formatter == "genero_plural":
        return _format_gender_ending(value, plural=True)
    if formatter == "genero_sustantivo":
        return _format_gender_noun_ending(value)
    if formatter == "genero_sustantivo_plural":
        return _format_gender_noun_ending(value, plural=True)
    logging.warning("Formato de marcador desconocido: %s", formatter)
    return value


def replace_placeholders(text: str, values: dict[str, str], missing: set[str]) -> str:
    def repl(match: re.Match[str]) -> str:
        key, formatter = _parse_placeholder(match.group(1))
        if key not in values:
            if key.casefold() == "foto":
                return ""
            missing.add(key)
            return match.group(0)
        return _format_placeholder_value(values.get(key, ""), formatter)

    return PLACEHOLDER_RE.sub(repl, text or "")


def _replace_paragraph(paragraph, values: dict[str, str], missing: set[str]) -> None:
    full_text = "".join(run.text for run in paragraph.runs)
    if "{{" not in full_text:
        return
    replaced = replace_placeholders(full_text, values, missing)
    if replaced == full_text:
        return
    if paragraph.runs:
        paragraph.runs[0].text = replaced
        for run in paragraph.runs[1:]:
            run.text = ""
    else:
        paragraph.add_run(replaced)


def _iter_paragraphs(container):
    for paragraph in getattr(container, "paragraphs", []):
        yield paragraph
    for table in getattr(container, "tables", []):
        for row in table.rows:
            for cell in row.cells:
                yield from _iter_paragraphs(cell)


def process_docx(path: Path, values: dict[str, str], missing: set[str]) -> None:
    document = Document(path)
    for paragraph in _iter_paragraphs(document):
        _replace_paragraph(paragraph, values, missing)
    for section in document.sections:
        for part in (section.header, section.footer):
            for paragraph in _iter_paragraphs(part):
                _replace_paragraph(paragraph, values, missing)
    document.save(path)


def process_xlsx(
    path: Path,
    values: dict[str, str],
    missing: set[str],
    photo_path: str | Path | None = None,
) -> None:
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    has_photo_placeholder = any(
                        _parse_placeholder(match.group(1))[0].casefold() == "foto"
                        for match in PLACEHOLDER_RE.finditer(cell.value)
                    )
                    cell.value = replace_placeholders(cell.value, values, missing)
                    if has_photo_placeholder and photo_path:
                        image = ExcelImage(str(photo_path))
                        image.width, image.height = _photo_dimensions(sheet, cell.coordinate)
                        image.anchor = cell.coordinate
                        sheet.add_image(image)
        for attr in ("oddHeader", "evenHeader", "firstHeader", "oddFooter", "evenFooter", "firstFooter"):
            header_footer = getattr(sheet, attr, None)
            if header_footer:
                for side in ("left", "center", "right"):
                    item = getattr(header_footer, side, None)
                    if item and item.text:
                        item.text = replace_placeholders(item.text, values, missing)
    workbook.save(path)


def _photo_dimensions(sheet, coordinate: str) -> tuple[int, int]:
    """Fit a photo to the merged area containing the Foto placeholder."""
    target = None
    for merged_range in sheet.merged_cells.ranges:
        if coordinate in merged_range:
            target = merged_range
            break
    if target is None:
        min_col = max_col = sheet[coordinate].column
        min_row = max_row = sheet[coordinate].row
    else:
        min_col, min_row, max_col, max_row = range_boundaries(str(target))

    width = 0.0
    for column in range(min_col, max_col + 1):
        letter = sheet.cell(row=1, column=column).column_letter
        width += float(sheet.column_dimensions[letter].width or 13.0) * 7 + 5

    height_points = 0.0
    for row in range(min_row, max_row + 1):
        height_points += float(sheet.row_dimensions[row].height or 15.0)

    width_pixels = max(1, int(width - 6))
    height_pixels = max(1, int(height_points * 96 / 72 - 6))
    return width_pixels, height_pixels


def process_template_copy(
    template: Path,
    destination: Path,
    values: dict[str, str],
    photo_path: str | Path | None = None,
) -> TemplateResult:
    result = TemplateResult()
    missing: set[str] = set()
    suffix = template.suffix.lower()
    if suffix in {".doc", ".xls"}:
        result.skipped_files.append(template.name)
        logging.warning(
            "Plantilla heredada omitida: %s. Conviertala a DOCX/XLSX o instale soporte de Microsoft Office.",
            template,
        )
        return result
    shutil.copy2(template, destination)
    if STATIC_TEMPLATE_MARKER in template.stem.casefold():
        logging.info("Plantilla estatica copiada sin reemplazo: %s", template)
        result.generated_files.append(destination)
        return result
    try:
        if suffix == ".docx":
            process_docx(destination, values, missing)
        elif suffix == ".xlsx":
            process_xlsx(destination, values, missing, photo_path=photo_path)
        else:
            result.skipped_files.append(template.name)
    except Exception:
        destination.unlink(missing_ok=True)
        raise
    result.generated_files.append(destination)
    result.missing_placeholders.update(missing)
    return result
