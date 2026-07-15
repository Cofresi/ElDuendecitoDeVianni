from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


POLICY_FIELD = "Política Horario"
DERIVED_FIELD = "Horario Laboral"


@dataclass(frozen=True)
class SchedulePolicy:
    horario1: str
    horario2: str
    days: int | None
    break_minutes: int | None
    holidays: bool


class WorkScheduleLookup:
    def __init__(self, policies: dict[str, SchedulePolicy]):
        self.policies = policies

    @classmethod
    def from_file(cls, path: str | Path) -> "WorkScheduleLookup":
        lookup_path = Path(path)
        if not lookup_path.exists():
            logging.warning("Tabla de horarios no encontrada: %s", lookup_path)
            return cls({})

        try:
            workbook = load_workbook(lookup_path, data_only=True, read_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        except Exception as exc:
            logging.warning("No se pudo leer la tabla de horarios %s: %s", lookup_path, exc)
            return cls({})

        if not rows:
            return cls({})

        headers = [_normalize_header(value) for value in rows[0]]
        columns = {
            "horario1": _find_column(headers, "horario1"),
            "horario2": _find_column(headers, "horario2"),
            "dias": _find_column(headers, "dias"),
            "break": _find_column(headers, "break"),
            "feriados": _find_column(headers, "feriados"),
        }
        if columns["horario1"] is None:
            logging.warning("La tabla de horarios no tiene columna horario1.")
            return cls({})

        policies: dict[str, SchedulePolicy] = {}
        for row in rows[1:]:
            key = _cell_text(row, columns["horario1"])
            if not key:
                continue
            policies[key.casefold()] = SchedulePolicy(
                horario1=key,
                horario2=_cell_text(row, columns["horario2"]),
                days=_cell_int(row, columns["dias"]),
                break_minutes=_cell_int(row, columns["break"]),
                holidays=_cell_bool(row, columns["feriados"]),
            )

        logging.info("Tabla de horarios cargada: %s politicas.", len(policies))
        return cls(policies)

    def sentence_for(self, policy_name: str) -> str:
        policy_key = policy_name.strip().casefold()
        if not policy_key:
            return ""

        policy = self.policies.get(policy_key)
        if not policy:
            logging.warning("Politica de horario sin coincidencia: %s", policy_name)
            return ""

        return build_work_schedule_sentence(policy)


def add_work_schedule_sentence(employee: dict[str, str], lookup: WorkScheduleLookup) -> None:
    employee[DERIVED_FIELD] = lookup.sentence_for(employee.get(POLICY_FIELD, ""))


def build_work_schedule_sentence(policy: SchedulePolicy) -> str:
    horario1 = _format_schedule(policy.horario1)
    horario2 = _format_schedule(policy.horario2)
    if not horario1:
        return ""

    if policy.days == 5:
        sentence = f"Lunes a Viernes de {horario1}"
        if horario2:
            sentence += f" y Sábado de {horario2}"
    elif policy.days == 6:
        sentence = f"Lunes a Sábado de {horario1}"
        if horario2:
            sentence += f" y Domingo de {horario2}"
    else:
        sentence = horario1

    if policy.break_minutes:
        sentence += f", con {policy.break_minutes} minutos de break"

    if policy.days == 6:
        connector = " y" if policy.break_minutes else ", con"
        sentence += f"{connector} un día libre a la semana según programación"

    if policy.holidays:
        sentence += ", este horario incluye los días feriados"

    return sentence + "."


def _format_schedule(value: str) -> str:
    text = re.sub(r"\s*\([^)]*\)", "", value).strip()
    text = re.sub(r"\s*\*+\s*$", "", text)
    text = re.sub(r"\s+[Aa]\s+", " a ", text)
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\b0(\d:\d{2})\s*([AP]M)\b", r"\1 \2", text, flags=re.IGNORECASE)
    text = re.sub(r"\b([AP])M\b", lambda match: match.group(0).lower(), text, flags=re.IGNORECASE)
    return text


def _normalize_header(value: object) -> str:
    return str(value or "").strip().casefold()


def _find_column(headers: list[str], expected: str) -> int | None:
    for index, header in enumerate(headers):
        if header == expected or header.startswith(expected):
            return index
    return None


def _cell_value(row: tuple[object, ...], index: int | None) -> object:
    if index is None or index >= len(row):
        return None
    return row[index]


def _cell_text(row: tuple[object, ...], index: int | None) -> str:
    value = _cell_value(row, index)
    if value is None:
        return ""
    return str(value).strip()


def _cell_int(row: tuple[object, ...], index: int | None) -> int | None:
    value = _cell_value(row, index)
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        logging.warning("Valor numerico invalido en tabla de horarios: %s", value)
        return None


def _cell_bool(row: tuple[object, ...], index: int | None) -> bool:
    value = _cell_value(row, index)
    if value in (None, ""):
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().casefold()
    return text in {"1", "1.0", "true", "si", "sí", "yes", "y"}
